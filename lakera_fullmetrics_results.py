import os
import argparse
import asyncio
import json
import time
from typing import List, Dict, Optional, Union, Tuple
from datetime import datetime
from glob import glob
from dataclasses import dataclass
import aiohttp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm
import pandas as pd
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix

@dataclass
class ConfigBase:
    lakera_api_key: Optional[str] = None
    LAKERA_RESULTS_URL: str = 'https://api.lakera.ai/v2/guard/results'

CATEGORIES = [
    'moderated_content/hate', 'moderated_content/profanity', 'moderated_content/sexual',
    'moderated_content/violence', 'moderated_content/weapons', 'moderated_content/crime',
    'pii/address', 'pii/credit_card', 'pii/email', 'pii/iban_code', 'pii/ip_address',
    'pii/name', 'pii/phone_number', 'pii/us_social_security_number', 'prompt_attack',
    'unknown_links','accuracy','precision','recall','f1 score','false positive rate', 
    'true positive rate','false negative rate','true negative rate', 'latency'
]

@dataclass
class ClassificationCounts:
    l1_confident: int = 0
    l2_very_likely: int = 0
    l3_likely: int = 0
    l4_less_likely: int = 0
    l5_unlikely: int = 0
    l6_latency: int = 0

@dataclass
class PerformanceMetrics:
    accuracy: float = 0.0
    precision: float = 0.0
    recall: float = 0.0
    f1_score: float = 0.0
    false_positive_rate: float = 0.0
    false_negative_rate: float = 0.0
    true_positive_rate: float = 0.0
    true_negative_rate: float = 0.0
    avg_latency: float = 0.0

class DatasetReader:
    """Handles reading different file formats and extracting text content"""
    
    @staticmethod
    def read_file(file_path: str, text_column: Optional[str] = None, label_column: Optional[str] = None) -> Tuple[List[str], List[Optional[bool]]]:
        """
        Read various file formats and extract text content and labels
        
        Args:
            file_path: Path to the dataset file
            text_column: Column name containing text (for structured formats)
            label_column: Column name containing ground truth labels (for performance metrics)
            
        Returns:
            Tuple of (text strings to evaluate, ground truth labels)
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_ext == '.txt':
                return DatasetReader._read_txt(file_path), [None] * DatasetReader._count_lines_txt(file_path)
            elif file_ext == '.csv':
                return DatasetReader._read_csv(file_path, text_column, label_column)
            elif file_ext == '.jsonl':
                return DatasetReader._read_jsonl(file_path, text_column, label_column)
            elif file_ext == '.parquet':
                return DatasetReader._read_parquet(file_path, text_column, label_column)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
                
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            return [], []
    
    @staticmethod
    def _count_lines_txt(file_path: str) -> int:
        """Count lines in text file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return sum(1 for line in f if line.strip())
    
    @staticmethod
    def _read_txt(file_path: str) -> List[str]:
        """Read plain text file (one prompt per line)"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]
    
    @staticmethod
    def _read_csv(file_path: str, text_column: Optional[str] = None, label_column: Optional[str] = None) -> Tuple[List[str], List[Optional[bool]]]:
        """Read CSV file and extract text column and labels"""
        df = pd.read_csv(file_path)
        
        # Auto-detect ground truth column if not specified
        if label_column is None:
            detected_label_col = DatasetReader.auto_detect_ground_truth_column(df)
            if detected_label_col:
                label_column = detected_label_col
                print(f"Auto-detected ground truth column: '{label_column}'")
        
        # Get text data
        if text_column:
            if text_column not in df.columns:
                raise ValueError(f"Column '{text_column}' not found in CSV. Available columns: {list(df.columns)}")
            texts = df[text_column].astype(str).tolist()
        else:
            # Auto-detect text column
            text_col = DatasetReader._auto_detect_text_column(df)
            if text_col:
                texts = df[text_col].astype(str).tolist()
                print(f"Auto-detected text column: '{text_col}'")
            else:
                # If only one column, use it
                if len(df.columns) == 1:
                    texts = df.iloc[:, 0].astype(str).tolist()
                else:
                    raise ValueError(f"Multiple columns found. Please specify --text-column. Available: {list(df.columns)}")
        
        # Get labels if specified or detected
        labels = [None] * len(texts)
        if label_column:
            if label_column in df.columns:
                labels = DatasetReader._parse_labels(df[label_column].tolist())
                valid_labels = sum(1 for l in labels if l is not None)
                print(f"Found {valid_labels} valid ground truth labels in column '{label_column}'")
            else:
                print(f"Warning: Label column '{label_column}' not found. Performance metrics will not be calculated.")
        
        return texts, labels
    
    @staticmethod
    def _read_jsonl(file_path: str, text_column: Optional[str] = None, label_column: Optional[str] = None) -> Tuple[List[str], List[Optional[bool]]]:
        """Read JSONL file and extract text content and labels"""
        prompts = []
        labels = []
        sample_data = None
        
        # First pass: read a sample to auto-detect columns if needed
        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                if line.strip() and line_num <= 5:  # Check first 5 lines for auto-detection
                    try:
                        data = json.loads(line.strip())
                        if sample_data is None:
                            sample_data = data
                        break
                    except json.JSONDecodeError:
                        continue
        
        # Auto-detect columns if not specified
        if sample_data:
            if text_column is None:
                text_column = DatasetReader._extract_text_field_name(sample_data)
                if text_column:
                    print(f"Auto-detected text field: '{text_column}'")
            
            if label_column is None:
                label_column = DatasetReader._extract_label_field_name(sample_data)
                if label_column:
                    print(f"Auto-detected ground truth field: '{label_column}'")
        
        # Second pass: extract data
        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                if line.strip():
                    try:
                        data = json.loads(line.strip())
                        
                        # Extract text
                        if text_column and text_column in data:
                            prompts.append(str(data[text_column]))
                        else:
                            # Auto-detect text field
                            text_value = DatasetReader._extract_text_from_json(data)
                            if text_value:
                                prompts.append(text_value)
                            else:
                                print(f"Warning: No suitable text field found in line {line_num}")
                                continue
                        
                        # Extract label
                        if label_column and label_column in data:
                            labels.append(DatasetReader._parse_single_label(data[label_column]))
                        else:
                            labels.append(None)
                                
                    except json.JSONDecodeError as e:
                        print(f"Warning: Invalid JSON on line {line_num}: {e}")
        
        if labels and any(l is not None for l in labels):
            valid_labels = sum(1 for l in labels if l is not None)
            print(f"Found {valid_labels} valid ground truth labels")
                        
        return prompts, labels
    
    @staticmethod
    def _read_parquet(file_path: str, text_column: Optional[str] = None, label_column: Optional[str] = None) -> Tuple[List[str], List[Optional[bool]]]:
        """Read Parquet file and extract text column and labels"""
        try:
            df = pd.read_parquet(file_path)
            
            # Auto-detect ground truth column if not specified
            if label_column is None:
                detected_label_col = DatasetReader.auto_detect_ground_truth_column(df)
                if detected_label_col:
                    label_column = detected_label_col
                    print(f"Auto-detected ground truth column: '{label_column}'")
            
            # Get text data
            if text_column:
                if text_column not in df.columns:
                    raise ValueError(f"Column '{text_column}' not found in Parquet. Available columns: {list(df.columns)}")
                texts = df[text_column].astype(str).tolist()
            else:
                # Auto-detect text column
                text_col = DatasetReader._auto_detect_text_column(df)
                if text_col:
                    texts = df[text_col].astype(str).tolist()
                    print(f"Auto-detected text column: '{text_col}'")
                else:
                    # If only one column, use it
                    if len(df.columns) == 1:
                        texts = df.iloc[:, 0].astype(str).tolist()
                    else:
                        raise ValueError(f"Multiple columns found. Please specify --text-column. Available: {list(df.columns)}")
            
            # Get labels if specified or detected
            labels = [None] * len(texts)
            if label_column:
                if label_column in df.columns:
                    labels = DatasetReader._parse_labels(df[label_column].tolist())
                    valid_labels = sum(1 for l in labels if l is not None)
                    print(f"Found {valid_labels} valid ground truth labels in column '{label_column}'")
                else:
                    print(f"Warning: Label column '{label_column}' not found. Performance metrics will not be calculated.")
            
            return texts, labels
            
        except ImportError:
            raise ImportError("pandas and pyarrow are required to read Parquet files. Install with: pip install pandas pyarrow")
    
    @staticmethod
    def _auto_detect_text_column(df: pd.DataFrame) -> Optional[str]:
        """Auto-detect the most likely text column in a DataFrame"""
        # Common text column names
        text_column_names = ['text', 'prompt', 'content', 'message', 'input', 'query', 'question']
        
        # Check for exact matches first
        for col_name in text_column_names:
            if col_name.lower() in [col.lower() for col in df.columns]:
                return next(col for col in df.columns if col.lower() == col_name.lower())
        
        # Check for partial matches
        for col_name in text_column_names:
            for col in df.columns:
                if col_name.lower() in col.lower():
                    return col
        
        # Find the column with the longest average text length
        text_columns = []
        for col in df.columns:
            if df[col].dtype == 'object':  # String-like columns
                avg_length = df[col].astype(str).str.len().mean()
                if avg_length > 10:  # Assume text columns have longer content
                    text_columns.append((col, avg_length))
        
        if text_columns:
            # Return column with longest average length
            return max(text_columns, key=lambda x: x[1])[0]
        
        return None

    @staticmethod
    def auto_detect_ground_truth_column(df: pd.DataFrame) -> Optional[str]:
        """Auto-detect ground truth/label column in a DataFrame"""
        # Common label column names
        label_column_names = [
            'label', 'labels', 'ground_truth', 'groundtruth', 'gt', 'truth',
            'target', 'class', 'category', 'classification', 'is_threat',
            'threat', 'safe', 'unsafe', 'harmful', 'malicious', 'toxic',
            'flag', 'flagged', 'blocked', 'allowed', 'positive', 'negative',
            'outcome', 'result', 'verdict', 'decision', 'risk', 'score'
        ]
        
        # Check for exact matches first (case insensitive)
        for label_name in label_column_names:
            for col in df.columns:
                if col.lower() == label_name.lower():
                    return col
        
        # Check for partial matches
        for label_name in label_column_names:
            for col in df.columns:
                if label_name.lower() in col.lower() or col.lower() in label_name.lower():
                    return col
        
        # Look for columns with boolean-like values
        for col in df.columns:
            unique_values = set(str(v).lower().strip() for v in df[col].dropna().unique())
            
            # Check for boolean-like patterns
            boolean_patterns = [
                {'true', 'false'},
                {'1', '0'},
                {'yes', 'no'},
                {'threat', 'safe'},
                {'unsafe', 'safe'},
                {'harmful', 'harmless'},
                {'positive', 'negative'},
                {'malicious', 'benign'},
                {'toxic', 'non-toxic'},
                {'flagged', 'allowed'}
            ]
            
            for pattern in boolean_patterns:
                if unique_values.issubset(pattern) or pattern.issubset(unique_values):
                    return col
        
        return None
    
    @staticmethod
    def _parse_labels(label_list: List) -> List[Optional[bool]]:
        """Parse various label formats to boolean"""
        parsed_labels = []
        for label in label_list:
            parsed_labels.append(DatasetReader._parse_single_label(label))
        return parsed_labels
    
    @staticmethod
    def _parse_single_label(label) -> Optional[bool]:
        """Parse a single label to boolean"""
        if pd.isna(label) or label is None:
            return None
        
        if isinstance(label, bool):
            return label
        
        if isinstance(label, (int, float)):
            return bool(label)
        
        if isinstance(label, str):
            label_lower = label.lower().strip()
            # Expanded list of positive/threat indicators
            if label_lower in ['true', 'yes', '1', 'positive', 'threat', 'unsafe', 'harmful', 'toxic', 'malicious', 'flagged', 'blocked']:
                return True
            elif label_lower in ['false', 'no', '0', 'negative', 'safe', 'benign', 'harmless', 'allowed', 'clean']:
                return False
        
        return None
    
    @staticmethod
    def _extract_text_from_json(data: dict) -> Optional[str]:
        """Extract text content from JSON object"""
        # Common text field names
        text_fields = ['text', 'prompt', 'content', 'message', 'input', 'query', 'question']
        
        # Check for exact matches
        for field in text_fields:
            if field in data:
                return str(data[field])
        
        # Check for partial matches
        for field in text_fields:
            for key in data.keys():
                if field.lower() in key.lower():
                    return str(data[key])
        
        # If no obvious text field, look for string values
        string_values = [(k, v) for k, v in data.items() if isinstance(v, str) and len(v) > 10]
        if string_values:
            # Return the longest string value
            return max(string_values, key=lambda x: len(x[1]))[1]
        
        return None
    
    @staticmethod
    def _extract_text_field_name(data: dict) -> Optional[str]:
        """Extract the field name most likely to contain text from JSON object"""
        text_fields = ['text', 'prompt', 'content', 'message', 'input', 'query', 'question']
        
        # Check for exact matches
        for field in text_fields:
            if field in data:
                return field
        
        # Check for partial matches
        for field in text_fields:
            for key in data.keys():
                if field.lower() in key.lower():
                    return key
        
        return None
    
    @staticmethod
    def _extract_label_field_name(data: dict) -> Optional[str]:
        """Extract the field name most likely to contain labels from JSON object"""
        label_fields = [
            'label', 'labels', 'ground_truth', 'groundtruth', 'gt', 'truth',
            'target', 'class', 'category', 'classification', 'is_threat',
            'threat', 'safe', 'unsafe', 'harmful', 'malicious', 'toxic',
            'flag', 'flagged', 'blocked', 'allowed', 'positive', 'negative'
        ]
        
        # Check for exact matches
        for field in label_fields:
            if field in data:
                return field
        
        # Check for partial matches
        for field in label_fields:
            for key in data.keys():
                if field.lower() in key.lower():
                    return key
        
        return None

class LakeraEvaluator:
    def __init__(self, api_key: str):
        self.config = ConfigBase(lakera_api_key=api_key)
        self.headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

    async def validate_key(self) -> bool:
        try:
            async with aiohttp.ClientSession() as session:
                body = {"messages": [{"role": "user", "content": "test"}]}
                async with session.post(self.config.LAKERA_RESULTS_URL, headers=self.headers, json=body) as response:
                    return response.status == 200
        except Exception as e:
            print(f"Lakera API key validation failed: {str(e)}")
            return False

    async def check_prompt(self, prompt: str) -> Dict:
        start_time = time.time()
        async with aiohttp.ClientSession() as session:
            body = {"messages": [{"role": "user", "content": prompt}]}
            async with session.post(self.config.LAKERA_RESULTS_URL, headers=self.headers, json=body) as response:
                results_data = await response.json()
                end_time = time.time()
                latency = end_time - start_time
                return {
                    "prompt": prompt,
                    "results": results_data.get("results", []),
                    "latency": latency
                }

    def calculate_performance_metrics(self, results: List[Dict], ground_truth: List[Optional[bool]]) -> Dict[str, PerformanceMetrics]:
        """
        Calculate performance metrics for each detector type
        
        Args:
            results: List of API results
            ground_truth: List of ground truth labels (True for threat, False for safe, None for unknown)
            
        Returns:
            Dictionary mapping detector types to their performance metrics
        """
        if not any(label is not None for label in ground_truth):
            print("No ground truth labels available. Performance metrics will not be calculated.")
            return {}
        
        # Organize predictions by detector type
        detector_predictions = {}
        latencies = []
        
        for i, result in enumerate(results):
            if 'latency' in result:
                latencies.append(result['latency'])
            
            for detector in result['results']:
                detector_type = detector['detector_type']
                if detector_type not in detector_predictions:
                    detector_predictions[detector_type] = []
                
                # Convert Lakera results to binary predictions
                # L1-L3 are considered threats, L4-L5 are considered safe
                prediction = detector['result'] in ['l1_confident', 'l2_very_likely', 'l3_likely']
                detector_predictions[detector_type].append(prediction)
        
        avg_latency = sum(latencies) / len(latencies) if latencies else 0.0
        
        # Calculate metrics for each detector
        metrics = {}
        
        for detector_type, predictions in detector_predictions.items():
            # Filter out None values from ground truth and corresponding predictions
            valid_indices = [i for i, gt in enumerate(ground_truth) if gt is not None and i < len(predictions)]
            
            if not valid_indices:
                continue
                
            valid_gt = [ground_truth[i] for i in valid_indices]
            valid_pred = [predictions[i] for i in valid_indices]
            
            # Calculate confusion matrix
            tn, fp, fn, tp = confusion_matrix(valid_gt, valid_pred).ravel()
            
            # Calculate metrics
            accuracy = accuracy_score(valid_gt, valid_pred)
            precision = precision_score(valid_gt, valid_pred, zero_division=0)
            recall = recall_score(valid_gt, valid_pred, zero_division=0)
            f1 = f1_score(valid_gt, valid_pred, zero_division=0)
            
            # Calculate rates
            fpr = fp / (fp + tn) if (fp + tn) > 0 else 0.0  # False Positive Rate
            fnr = fn / (fn + tp) if (fn + tp) > 0 else 0.0  # False Negative Rate
            tpr = tp / (tp + fn) if (tp + fn) > 0 else 0.0  # True Positive Rate (Recall)
            tnr = tn / (tn + fp) if (tn + fp) > 0 else 0.0  # True Negative Rate (Specificity)
            
            metrics[detector_type] = PerformanceMetrics(
                accuracy=accuracy,
                precision=precision,
                recall=recall,
                f1_score=f1,
                false_positive_rate=fpr,
                false_negative_rate=fnr,
                true_positive_rate=tpr,
                true_negative_rate=tnr,
                avg_latency=avg_latency
            )
        
        return metrics

    def generate_classification_counts(self, results: List[Dict]) -> Dict[str, ClassificationCounts]:
        classifications = {}
        
        for result in results:
            for detector in result['results']:
                if detector['detector_type'] not in classifications:
                    classifications[detector['detector_type']] = ClassificationCounts()
                if hasattr(classifications[detector['detector_type']], detector['result']):
                    setattr(classifications[detector['detector_type']], detector['result'],
                           getattr(classifications[detector['detector_type']], detector['result']) + 1)
        
        return classifications

    def save_results(self, results: List[Dict], output_path: str, ground_truth: List[Optional[bool]] = None):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.dirname(output_path)
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        
        os.makedirs(output_dir, exist_ok=True)
        # Use dataset name for the Excel file
        filename = os.path.join(output_dir, f"{base_name}_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Style definitions
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        metric_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Get classification counts and performance metrics
        classifications = self.generate_classification_counts(results)
        performance_metrics = self.calculate_performance_metrics(results, ground_truth) if ground_truth else {}
        
        # Performance Metrics Sheet (First sheet if metrics available)
        if performance_metrics:
            metrics_sheet = wb.active
            metrics_sheet.title = "Performance Metrics"
            
            # Headers
            metrics_sheet.append([
                "Detector Type", "Accuracy (%)", "Precision", "Recall", "F1 Score",
                "False Positive Rate", "False Negative Rate", "True Positive Rate",
                "True Negative Rate", "Avg Latency (ms)"
            ])
            
            # Apply header styling
            for cell in metrics_sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Add metrics data
            for detector_type, metrics in performance_metrics.items():
                metrics_sheet.append([
                    detector_type,
                    f"{metrics.accuracy * 100:.2f}%",
                    f"{metrics.precision:.4f}",
                    f"{metrics.recall:.4f}",
                    f"{metrics.f1_score:.4f}",
                    f"{metrics.false_positive_rate:.4f}",
                    f"{metrics.false_negative_rate:.4f}",
                    f"{metrics.true_positive_rate:.4f}",
                    f"{metrics.true_negative_rate:.4f}",
                    f"{metrics.avg_latency * 1000:.2f}"
                ])
            
            # Add summary row
            if performance_metrics:
                metrics_sheet.append([])
                metrics_sheet.append(["OVERALL AVERAGES"])
                
                avg_accuracy = sum(m.accuracy for m in performance_metrics.values()) / len(performance_metrics)
                avg_precision = sum(m.precision for m in performance_metrics.values()) / len(performance_metrics)
                avg_recall = sum(m.recall for m in performance_metrics.values()) / len(performance_metrics)
                avg_f1 = sum(m.f1_score for m in performance_metrics.values()) / len(performance_metrics)
                avg_fpr = sum(m.false_positive_rate for m in performance_metrics.values()) / len(performance_metrics)
                avg_fnr = sum(m.false_negative_rate for m in performance_metrics.values()) / len(performance_metrics)
                avg_tpr = sum(m.true_positive_rate for m in performance_metrics.values()) / len(performance_metrics)
                avg_tnr = sum(m.true_negative_rate for m in performance_metrics.values()) / len(performance_metrics)
                avg_latency = list(performance_metrics.values())[0].avg_latency if performance_metrics else 0
                
                metrics_sheet.append([
                    "Average",
                    f"{avg_accuracy * 100:.2f}%",
                    f"{avg_precision:.4f}",
                    f"{avg_recall:.4f}",
                    f"{avg_f1:.4f}",
                    f"{avg_fpr:.4f}",
                    f"{avg_fnr:.4f}",
                    f"{avg_tpr:.4f}",
                    f"{avg_tnr:.4f}",
                    f"{avg_latency * 1000:.2f}"
                ])
        
        # Classification Table Sheet
        table_sheet = wb.create_sheet("Classification Table") if performance_metrics else wb.active
        if not performance_metrics:
            table_sheet.title = "Classification Table"
        
        # Count prompts with threats and no threats
        total_prompts = len(results)
        no_threats_count = sum(1 for result in results 
                             if all(detector['result'] == 'l5_unlikely' 
                                   for detector in result['results']))
        threats_count = total_prompts - no_threats_count

        # Add Threats Summary section
        table_sheet.append(["Summary"])
        table_sheet.append(["Category", "Count", "Percentage"])
        table_sheet.append(["Threats Detected", f"{threats_count}/{total_prompts}", f"{(threats_count/total_prompts)*100:.1f}%"])
        table_sheet.append(["No Threats", f"{no_threats_count}/{total_prompts}", f"{(no_threats_count/total_prompts)*100:.1f}%"])
        
        # Add blank rows for separation
        table_sheet.append([])
        table_sheet.append([])
        
        # Main classification table
        table_sheet.append(["Classification Breakdown"])
        table_sheet.append(["Category", "Confident", "Very Likely", "Likely", "Less Likely"])
        
        # Group the classifications into high-level categories
        grouped_classifications = {
            "Prompt Attack": ["prompt_attack"],
            "Content Safety": [cat for cat in CATEGORIES if cat.startswith("moderated_content/")],
            "Data Leak": [cat for cat in CATEGORIES if cat.startswith("pii/")],
            "Unknown Links": ["unknown_links"]
        }
        
        # Calculate totals for each group
        for group_name, categories in grouped_classifications.items():
            confident = sum(classifications.get(cat, ClassificationCounts()).l1_confident for cat in categories)
            very_likely = sum(classifications.get(cat, ClassificationCounts()).l2_very_likely for cat in categories)
            likely = sum(classifications.get(cat, ClassificationCounts()).l3_likely for cat in categories)
            less_likely = sum(classifications.get(cat, ClassificationCounts()).l4_less_likely for cat in categories)
            
            table_sheet.append([group_name, confident, very_likely, likely, less_likely])
        
        # Classification Breakdown Sheet
        breakdown_sheet = wb.create_sheet("Classification Breakdown")
        breakdown_sheet.append(["Classification Breakdown"])
        breakdown_sheet.append(["Detection Type", "L1 Confident", "L2 Very Likely", "L3 Likely", "L4 Less Likely", "L5 Unlikely"])
        
        for type_, counts in classifications.items():
            breakdown_sheet.append([
                type_, counts.l1_confident, counts.l2_very_likely, counts.l3_likely,
                counts.l4_less_likely, counts.l5_unlikely
            ])
        
        # Detailed Results Sheet
        results_sheet = wb.create_sheet("Detailed Results")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        header = ['Prompt']
        if ground_truth and any(gt is not None for gt in ground_truth):
            header.append('Ground Truth')
        header.append('Latency (ms)')
        for category in CATEGORIES:
            header.append(category)
        results_sheet.append(header)
        
        for i, result in enumerate(results):
            row = [result['prompt']]
            
            # Add ground truth if available
            if ground_truth and any(gt is not None for gt in ground_truth):
                if i < len(ground_truth) and ground_truth[i] is not None:
                    row.append('Threat' if ground_truth[i] else 'Safe')
                else:
                    row.append('Unknown')
            
            # Add latency
            latency_ms = result.get('latency', 0) * 1000
            row.append(f"{latency_ms:.2f}")
            
            for category in CATEGORIES:
                detector = next((d for d in result['results']
                               if d['detector_type'] == category), None)
                row.append(detector['result'] if detector else '')
            
            row_number = results_sheet.max_row + 1
            results_sheet.append(row)
            
            # If all results are L5, highlight the entire row in yellow
            if all(detector['result'] == 'l5_unlikely' for detector in result['results']):
                for cell in results_sheet[row_number]:
                    cell.fill = yellow_fill
        
        # Style headers and adjust column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
            
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        max_length = min(len(str(cell.value)), 50)
                    except:
                        pass
                ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        wb.save(filename)
        print(f"Results saved to {filename}")

def count_lines(file_path: str) -> int:
    """Count lines/rows in various file formats - improved robustness"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return sum(1 for line in f if line.strip())
        elif file_ext == '.csv':
            try:
                df = pd.read_csv(file_path)
                return len(df)
            except:
                # Fallback to line counting if pandas fails
                with open(file_path, 'r', encoding='utf-8') as f:
                    return sum(1 for line in f if line.strip()) - 1  # Subtract header
        elif file_ext == '.jsonl':
            with open(file_path, 'r', encoding='utf-8') as f:
                return sum(1 for line in f if line.strip())
        elif file_ext == '.parquet':
            try:
                df = pd.read_parquet(file_path)
                return len(df)
            except:
                return 0
        else:
            return 0
    except Exception as e:
        print(f"Warning: Could not count lines in {file_path}: {e}")
        return 0

def find_dataset_files(dataset_paths: List[str]) -> List[str]:
    """Find and validate dataset files from input paths or patterns - improved robustness"""
    found_files = []
    
    for path_pattern in dataset_paths:
        original_pattern = path_pattern
        files_found_for_pattern = []
        
        # Strategy 1: Check if it's a glob pattern
        if '*' in path_pattern or '?' in path_pattern:
            matched_files = glob(path_pattern)
            files_found_for_pattern.extend(matched_files)
            
            # Also try in datasets directory
            dataset_pattern = os.path.join('datasets', path_pattern)
            matched_files = glob(dataset_pattern)
            files_found_for_pattern.extend(matched_files)
        
        # Strategy 2: Direct file path - try multiple variations
        else:
            # List of possible paths to try
            possible_paths = [
                path_pattern,  # Exact path as given
                os.path.join('datasets', path_pattern),  # In datasets directory
            ]
            
            # Try common extensions if no extension provided
            if not os.path.splitext(path_pattern)[1]:
                extensions = ['.txt', '.csv', '.jsonl', '.parquet']
                for ext in extensions:
                    possible_paths.extend([
                        path_pattern + ext,
                        os.path.join('datasets', path_pattern + ext)
                    ])
            
            # Check each possible path
            for possible_path in possible_paths:
                if os.path.exists(possible_path):
                    files_found_for_pattern.append(possible_path)
                    break  # Take the first match
            
            # If still not found, try glob patterns with the name
            if not files_found_for_pattern:
                # Try as a partial filename match
                glob_patterns = [
                    f"*{path_pattern}*",
                    f"datasets/*{path_pattern}*",
                    f"{path_pattern}*",
                    f"datasets/{path_pattern}*"
                ]
                
                for pattern in glob_patterns:
                    matched_files = glob(pattern)
                    if matched_files:
                        files_found_for_pattern.extend(matched_files)
                        break
        
        if files_found_for_pattern:
            # Remove duplicates while preserving order
            for file_path in files_found_for_pattern:
                if file_path not in found_files:
                    found_files.append(file_path)
        else:
            print(f"Warning: No files found for pattern/name '{original_pattern}'")
            # Try to give helpful suggestions
            dataset_dir = 'datasets'
            if os.path.exists(dataset_dir):
                all_files = os.listdir(dataset_dir)
                similar_files = [f for f in all_files if original_pattern.lower() in f.lower()]
                if similar_files:
                    print(f"  Did you mean one of these files in datasets/: {similar_files[:3]}")
    
    return found_files

async def evaluate_single_dataset(file_path: str, api_key: str, evaluator: LakeraEvaluator, text_column: Optional[str] = None, label_column: Optional[str] = None):
    """Evaluate a single dataset file"""
    try:
        prompts, ground_truth = DatasetReader.read_file(file_path, text_column, label_column)
        if not prompts:
            print(f"No valid prompts found in {file_path}")
            return False
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return False
    
    print(f"\nEvaluating dataset: {file_path}")
    print(f"Total prompts: {len(prompts)}")
    
    # Check if ground truth labels are available
    has_labels = any(label is not None for label in ground_truth)
    if has_labels:
        label_count = sum(1 for label in ground_truth if label is not None)
        threat_count = sum(1 for label in ground_truth if label is True)
        safe_count = sum(1 for label in ground_truth if label is False)
        print(f"Ground truth labels: {label_count} total ({threat_count} threats, {safe_count} safe)")
    else:
        print("No ground truth labels found - performance metrics will not be calculated")
    
    results = []
    with tqdm(total=len(prompts), desc=f"Processing {os.path.basename(file_path)}") as pbar:
        for prompt in prompts:
            try:
                result = await evaluator.check_prompt(prompt)
                results.append(result)
                await asyncio.sleep(0.1)  # Rate limiting
            except Exception as e:
                print(f"\nError processing prompt: {e}")
            pbar.update(1)
    
    # Use the original file path structure for output
    if file_path.startswith('datasets/'):
        results_path = os.path.join('results', file_path[9:])  # Remove 'datasets/' prefix
    else:
        results_path = os.path.join('results', os.path.basename(file_path))
    
    evaluator.save_results(results, results_path, ground_truth)
    return True

async def evaluate_multiple_datasets(dataset_paths: List[str], api_key: str, text_column: Optional[str] = None, label_column: Optional[str] = None):
    """Evaluate multiple datasets sequentially"""
    evaluator = LakeraEvaluator(api_key)
    
    # Validate API key once
    print("Validating Lakera API key...")
    if not await evaluator.validate_key():
        print("API key validation failed. Please check your API key.")
        return
    print("API key validated successfully!")
    
    # Find all dataset files
    dataset_files = find_dataset_files(dataset_paths)
    
    if not dataset_files:
        print("No valid dataset files found!")
        return
    
    print(f"\nFound {len(dataset_files)} dataset file(s) to process:")
    for i, file_path in enumerate(dataset_files, 1):
        line_count = count_lines(file_path)
        file_ext = os.path.splitext(file_path)[1].lower()
        print(f"  {i}. {file_path} ({line_count} rows) [{file_ext}]")
    
    # Process each file sequentially
    successful_count = 0
    total_files = len(dataset_files)
    
    for i, file_path in enumerate(dataset_files, 1):
        print(f"\n{'='*60}")
        print(f"Processing file {i}/{total_files}: {file_path}")
        print('='*60)
        
        success = await evaluate_single_dataset(file_path, api_key, evaluator, text_column, label_column)
        if success:
            successful_count += 1
        
        # Add a small delay between files to be respectful to the API
        if i < total_files:
            await asyncio.sleep(1)
    
    print(f"\n{'='*60}")
    print(f"BATCH PROCESSING COMPLETE")
    print(f"Successfully processed: {successful_count}/{total_files} files")
    print('='*60)

def main():
    parser = argparse.ArgumentParser(
        description='Lakera Guard Evaluation Tool - Multiple File Format Support with Performance Metrics',
        epilog="""
Supported file formats:
  - .txt (one prompt per line)
  - .csv (comma-separated values)
  - .jsonl (JSON Lines format)
  - .parquet (Apache Parquet format)

Examples:
  # Process files with robust discovery (like original script)
  python extended_lakera_automated_results.py --datasets mini_JailBreakV_28K
  
  # Process specific files with extensions
  python extended_lakera_automated_results.py --datasets file1.txt file2.csv data/file3.jsonl
  
  # Process all CSV files in a directory
  python extended_lakera_automated_results.py --datasets "data/*.csv"
  
  # Process with auto-detected ground truth labels
  python extended_lakera_automated_results.py --datasets safety_data.parquet --text-column "prompt"
  
  # Process with specified ground truth column
  python extended_lakera_automated_results.py --datasets labeled_data.csv --text-column "text" --label-column "is_threat"
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument('--datasets', nargs='+', required=True,
                       help='Paths to dataset files or patterns. Supports partial names, auto-extension detection, and glob patterns.')
    parser.add_argument('--text-column', 
                       help='Column name containing text for structured formats. Will auto-detect if not specified.')
    parser.add_argument('--label-column', 
                       help='Column name containing ground truth labels. Will auto-detect common label columns if not specified.')
    parser.add_argument('-e', '--env', 
                       help='Lakera Guard API key (alternatively, set LAKERA_GUARD_API_KEY environment variable)')
    
    args = parser.parse_args()
    
    api_key = args.env or os.getenv('LAKERA_GUARD_API_KEY')
    if not api_key:
        print("\nError: Lakera Guard API key not found!")
        print("\nYou can provide the API key in one of two ways:")
        print("1. Set the LAKERA_GUARD_API_KEY environment variable:")
        print("   export LAKERA_GUARD_API_KEY=your_api_key")
        print("\n2. Use the -e/--env flag:")
        print("   python extended_lakera_automated_results.py --datasets file1.txt -e your_api_key")
        return
    
    # Check for required dependencies based on file types
    dataset_files = find_dataset_files(args.datasets)
    
    if not dataset_files:
        print("\nNo valid dataset files found!")
        print("\nTips for file discovery:")
        print("- Files can be specified without extensions (will try .txt, .csv, .jsonl, .parquet)")
        print("- Files are searched in current directory and 'datasets/' subdirectory")
        print("- Partial filenames are supported (e.g., 'mini_Jail' will match 'mini_JailBreakV_28K.txt')")
        print("- Use glob patterns for multiple files (e.g., '*.csv')")
        return
    
    file_extensions = {os.path.splitext(f)[1].lower() for f in dataset_files}
    
    missing_deps = []
    if '.parquet' in file_extensions:
        try:
            import pyarrow
        except ImportError:
            missing_deps.append("pyarrow")
    
    if '.csv' in file_extensions or '.parquet' in file_extensions:
        try:
            import pandas
        except ImportError:
            missing_deps.append("pandas")
    
    if missing_deps:
        print(f"\nError: Missing required dependencies: {', '.join(missing_deps)}")
        print(f"Install with: pip install {' '.join(missing_deps)}")
        return
    
    # Check for scikit-learn dependency for performance metrics
    try:
        import sklearn
    except ImportError:
        print("\nWarning: scikit-learn not found. Performance metrics will not be calculated.")
        print("Install with: pip install scikit-learn")
    
    asyncio.run(evaluate_multiple_datasets(args.datasets, api_key, args.text_column, args.label_column))

if __name__ == '__main__':
    main()