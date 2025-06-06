import os
import argparse
import asyncio
from typing import List, Dict, Optional
from datetime import datetime
from glob import glob
from dataclasses import dataclass
import aiohttp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm

@dataclass
class ConfigBase:
    lakera_api_key: Optional[str] = None
    LAKERA_RESULTS_URL: str = 'https://api.lakera.ai/v2/guard/results'

CATEGORIES = [
    'moderated_content/hate', 'moderated_content/profanity', 'moderated_content/sexual',
    'moderated_content/violence', 'moderated_content/weapons', 'moderated_content/crime',
    'pii/address', 'pii/credit_card', 'pii/email', 'pii/iban_code', 'pii/ip_address',
    'pii/name', 'pii/phone_number', 'pii/us_social_security_number', 'prompt_attack',
    'unknown_links','accuracy','precision','recall','f1 score','false positive rate', 'true positive rate','false negative rate','true negative rate', 'latency'
]

@dataclass
class ClassificationCounts:
    l1_confident: int = 0
    l2_very_likely: int = 0
    l3_likely: int = 0
    l4_less_likely: int = 0
    l5_unlikely: int = 0
    l6_latency: int = 0
    
class LakeraEvaluator:
    def __init__(self, api_key: str):
        self.config = ConfigBase(lakera_api_key=api_key)
        self.headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

    async def validate_key(self) -> bool:
        try:
            async with aiohttp.ClientSession() as session:
                body = {"messages": [{"role": "user", "content": "test"}]}
                async with session.post(self.config.LAKERA_RESULTS_URL, headers=self.headers, json=body) as response:
                    return response.status == 200
        except Exception as e:
            print(f"Lakera API key validation failed: {str(e)}")
            return False

    async def check_prompt(self, prompt: str) -> Dict:
        async with aiohttp.ClientSession() as session:
            body = {"messages": [{"role": "user", "content": prompt}]}
            async with session.post(self.config.LAKERA_RESULTS_URL, headers=self.headers, json=body) as response:
                results_data = await response.json()
                return {
                    "prompt": prompt,
                    "results": results_data.get("results", [])
                }

    def generate_classification_counts(self, results: List[Dict]) -> Dict[str, ClassificationCounts]:
        classifications = {}
        
        for result in results:
            for detector in result['results']:
                if detector['detector_type'] not in classifications:
                    classifications[detector['detector_type']] = ClassificationCounts()
                if hasattr(classifications[detector['detector_type']], detector['result']):
                    setattr(classifications[detector['detector_type']], detector['result'],
                           getattr(classifications[detector['detector_type']], detector['result']) + 1)
        
        return classifications

    def save_results(self, results: List[Dict], output_path: str):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.dirname(output_path)
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        
        os.makedirs(output_dir, exist_ok=True)
        # Use dataset name for the Excel file
        filename = os.path.join(output_dir, f"{base_name}_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Style definitions
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        
        # Get classification counts
        classifications = self.generate_classification_counts(results)
        
        # Classification Table Sheet (First sheet)
        table_sheet = wb.active
        table_sheet.title = "Classification Table"
        
        # Count prompts with threats and no threats
        total_prompts = len(results)
        no_threats_count = sum(1 for result in results 
                             if all(detector['result'] == 'l5_unlikely' 
                                   for detector in result['results']))
        threats_count = total_prompts - no_threats_count

        # Add Threats Summary section
        table_sheet.append(["Summary"])
        table_sheet.append(["Category", "Count", "Percentage"])
        table_sheet.append(["Threats Detected", f"{threats_count}/{total_prompts}", f"{(threats_count/total_prompts)*100:.1f}%"])
        table_sheet.append(["No Threats", f"{no_threats_count}/{total_prompts}", f"{(no_threats_count/total_prompts)*100:.1f}%"])
        
        # Add blank rows for separation
        table_sheet.append([])
        table_sheet.append([])
        
        # Main classification table
        table_sheet.append(["Classification Breakdown"])
        table_sheet.append(["Category", "Confident", "Very Likely", "Likely", "Less Likely"])
        
        # Group the classifications into high-level categories
        grouped_classifications = {
            "Prompt Attack": ["prompt_attack"],
            "Content Safety": [cat for cat in CATEGORIES if cat.startswith("moderated_content/")],
            "Data Leak": [cat for cat in CATEGORIES if cat.startswith("pii/")],
            "Unknown Links": ["unknown_links"]
        }
        
        # Calculate totals for each group
        for group_name, categories in grouped_classifications.items():
            confident = sum(classifications.get(cat, ClassificationCounts()).l1_confident for cat in categories)
            very_likely = sum(classifications.get(cat, ClassificationCounts()).l2_very_likely for cat in categories)
            likely = sum(classifications.get(cat, ClassificationCounts()).l3_likely for cat in categories)
            less_likely = sum(classifications.get(cat, ClassificationCounts()).l4_less_likely for cat in categories)
            
            table_sheet.append([group_name, confident, very_likely, likely, less_likely])
        
        # Classification Breakdown Sheet
        breakdown_sheet = wb.create_sheet("Classification Breakdown")
        breakdown_sheet.append(["Classification Breakdown"])
        breakdown_sheet.append(["Detection Type", "L1 Confident", "L2 Very Likely", "L3 Likely", "L4 Less Likely", "L5 Unlikely"])
        
        for type_, counts in classifications.items():
            breakdown_sheet.append([
                type_, counts.l1_confident, counts.l2_very_likely, counts.l3_likely,
                counts.l4_less_likely, counts.l5_unlikely
            ])
        
        # Detailed Results Sheet
        results_sheet = wb.create_sheet("Detailed Results")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        header = ['Prompt']
        for category in CATEGORIES:
            header.append(category)
        results_sheet.append(header)
        
        for result in results:
            row = [result['prompt']]
            
            for category in CATEGORIES:
                detector = next((d for d in result['results']
                               if d['detector_type'] == category), None)
                row.append(detector['result'] if detector else '')
            
            row_number = results_sheet.max_row + 1
            results_sheet.append(row)
            
            # If all results are L5, highlight the entire row in yellow
            if all(detector['result'] == 'l5_unlikely' for detector in result['results']):
                for cell in results_sheet[row_number]:
                    cell.fill = yellow_fill
        
        # Style headers and adjust column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
            
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        max_length = min(len(str(cell.value)), 50)
                    except:
                        pass
                ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        wb.save(filename)
        print(f"Results saved to {filename}")

def count_lines(file_path: str) -> int:
    try:
        with open(file_path, 'r') as f:
            return sum(1 for line in f if line.strip())
    except Exception:
        return 0

def find_dataset_files(dataset_paths: List[str]) -> List[str]:
    """Find and validate dataset files from input paths or patterns"""
    found_files = []
    
    for path_pattern in dataset_paths:
        # Check if it's a glob pattern or direct file path
        if '*' in path_pattern or '?' in path_pattern:
            # It's a glob pattern
            matched_files = glob(path_pattern)
            if not matched_files:
                print(f"Warning: No files found matching pattern '{path_pattern}'")
            else:
                found_files.extend(matched_files)
        else:
            # Direct file path - check relative to datasets dir first
            relative_path = os.path.join('datasets', path_pattern)
            if os.path.exists(relative_path):
                found_files.append(relative_path)
            elif os.path.exists(path_pattern):
                found_files.append(path_pattern)
            else:
                print(f"Warning: Dataset file '{path_pattern}' not found")
    
    return found_files

async def evaluate_single_dataset(file_path: str, api_key: str, evaluator: LakeraEvaluator):
    """Evaluate a single dataset file"""
    try:
        with open(file_path, 'r') as f:
            prompts = [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return False
    
    print(f"\nEvaluating dataset: {file_path}")
    print(f"Total prompts: {len(prompts)}")
    
    results = []
    with tqdm(total=len(prompts), desc=f"Processing {os.path.basename(file_path)}") as pbar:
        for prompt in prompts:
            try:
                result = await evaluator.check_prompt(prompt)
                results.append(result)
                await asyncio.sleep(0.1)  # Rate limiting
            except Exception as e:
                print(f"\nError processing prompt: {e}")
            pbar.update(1)
    
    # Use the original file path structure for output
    if file_path.startswith('datasets/'):
        results_path = os.path.join('results', file_path[9:])  # Remove 'datasets/' prefix
    else:
        results_path = os.path.join('results', os.path.basename(file_path))
    
    evaluator.save_results(results, results_path)
    return True

async def evaluate_multiple_datasets(dataset_paths: List[str], api_key: str):
    """Evaluate multiple datasets sequentially"""
    evaluator = LakeraEvaluator(api_key)
    
    # Validate API key once
    print("Validating Lakera API key...")
    if not await evaluator.validate_key():
        print("API key validation failed. Please check your API key.")
        return
    print("API key validated successfully!")
    
    # Find all dataset files
    dataset_files = find_dataset_files(dataset_paths)
    
    if not dataset_files:
        print("No valid dataset files found!")
        return
    
    print(f"\nFound {len(dataset_files)} dataset file(s) to process:")
    for i, file_path in enumerate(dataset_files, 1):
        line_count = count_lines(file_path)
        print(f"  {i}. {file_path} ({line_count} lines)")
    
    # Process each file sequentially
    successful_count = 0
    total_files = len(dataset_files)
    
    for i, file_path in enumerate(dataset_files, 1):
        print(f"\n{'='*60}")
        print(f"Processing file {i}/{total_files}: {file_path}")
        print('='*60)
        
        success = await evaluate_single_dataset(file_path, api_key, evaluator)
        if success:
            successful_count += 1
        
        # Add a small delay between files to be respectful to the API
        if i < total_files:
            await asyncio.sleep(1)
    
    print(f"\n{'='*60}")
    print(f"BATCH PROCESSING COMPLETE")
    print(f"Successfully processed: {successful_count}/{total_files} files")
    print('='*60)

def main():
    parser = argparse.ArgumentParser(
        description='Lakera Guard Evaluation Tool - Multiple File Support',
        epilog="""
Examples:
  # Process specific files
  python lakera_results.py --datasets file1.txt file2.csv data/file3.txt
  
  # Process all CSV files in a directory
  python lakera_results.py --datasets "data/*.csv"
  
  # Process all txt files in datasets directory
  python lakera_results.py --datasets "datasets/*.txt"
  
  # Mix of specific files and patterns
  python lakera_results.py --datasets file1.txt "data/*.csv" specific_file.txt
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Change from --dataset to --datasets and allow multiple values
    parser.add_argument('--datasets', nargs='+', required=True,
                       help='Paths to dataset files or glob patterns (e.g., "*.csv" "data/*.txt")')
    parser.add_argument('-e', '--env', 
                       help='Lakera Guard API key (alternatively, set LAKERA_GUARD_API_KEY environment variable)')
    
    args = parser.parse_args()
    
    api_key = args.env or os.getenv('LAKERA_GUARD_API_KEY')
    if not api_key:
        print("\nError: Lakera Guard API key not found!")
        print("\nYou can provide the API key in one of two ways:")
        print("1. Set the LAKERA_GUARD_API_KEY environment variable:")
        print("   export LAKERA_GUARD_API_KEY=your_api_key")
        print("\n2. Use the -e/--env flag:")
        print("   python lakera_results.py --datasets file1.txt file2.txt -e your_api_key")
        return
    
    asyncio.run(evaluate_multiple_datasets(args.datasets, api_key))

if __name__ == '__main__':
    main()
